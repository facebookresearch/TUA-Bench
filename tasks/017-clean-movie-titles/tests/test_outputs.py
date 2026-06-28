# Copyright (c) Meta Platforms, Inc. and affiliates.
# All rights reserved.
#
# This source code is licensed under the license found in the
# LICENSE file in the root directory of this source tree.

TASK_SPEC = {'example_id': 'a9f325aa-8c05-4e4f-8341-9e4358565f4f',
 'inputs': [{'dest': 'Movie_title_garbage_clean.xlsx',
             'url': 'https://huggingface.co/datasets/xlangai/ubuntu_osworld_file_cache/resolve/main/libreoffice_calc/a9f325aa-8c05-4e4f-8341-9e4358565f4f/Movie_titles_garbage_clean.xlsx'}],
 'instruction': "I want to copy the movie titles from the 'Garbage Movie Titles' column to the "
                "'Clean Movie Titles' column. Please remove redundant whitespace and canonicalize "
                'letter casing by capitalizing the first letter of each word and leaving the '
                "other letters lowercase. Finish the work and don't touch irrelevant regions, "
                'even if they are blank.',
 'metric_spec': {'expected': {'dest': 'Movie_title_garbage_clean_gold.xlsx',
                              'type': 'cloud_file',
                              'url': 'https://huggingface.co/datasets/xlangai/ubuntu_osworld_file_cache/resolve/main/libreoffice_calc/a9f325aa-8c05-4e4f-8341-9e4358565f4f/Movie_titles_garbage_clean_gold.xlsx'},
                 'func': 'compare_table',
                 'options': {'rules': [{'sheet_idx0': 0,
                                        'sheet_idx1': 'EI0',
                                        'type': 'sheet_data'}]},
                 'result': '/app/Movie_title_garbage_clean.xlsx'},
 'oracle': {'mode': 'fetch_copy',
            'url': 'https://huggingface.co/datasets/xlangai/ubuntu_osworld_file_cache/resolve/main/libreoffice_calc/a9f325aa-8c05-4e4f-8341-9e4358565f4f/Movie_titles_garbage_clean_gold.xlsx'},
 'output_path': '/app/Movie_title_garbage_clean.xlsx',
 'result_sidecars': [],
 'slug': '017-clean-movie-titles'}

import functools
import itertools
import json
import logging
import operator
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
import zipfile
from numbers import Number
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Match, Optional, Pattern, Tuple, TypeVar, Union, cast

import fitz
import formulas
import lxml.cssselect
import lxml.etree
import openpyxl
import pandas as pd
import xmltodict
from lxml.etree import _Element
from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.chart._chart import ChartBase
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.pivot.cache import CacheSource as PivotCacheSource
from openpyxl.pivot.table import TableDefinition as PivotTableDefinition
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.worksheet.filters import AutoFilter, SortState
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader
from rapidfuzz import fuzz

logger = logging.getLogger("tua.calc.osworld")
logging.basicConfig(level=logging.INFO)

LIMITATION_PHRASES = [
    "does not",
    "doesn't",
    "cannot",
    "can't",
    "not possible",
    "infeasible",
    "missing",
    "without",
    "unsupported",
    "unavailable",
    "not available",
    "not offered",
    "not listed",
    "nonexistent",
]

CACHE_DIR = Path("/tmp") / TASK_SPEC["slug"]
V = TypeVar("Value")


def get_app_dir() -> Path:
    return Path(os.environ.get("APP_DIR", "/app"))


def localize_app_path(path: str) -> Path:
    if path.startswith("/app/"):
        return get_app_dir() / path.removeprefix("/app/")
    return Path(path)


def fetch_cloud_file(url: str, dest: str) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    target = CACHE_DIR / dest
    if not target.exists():
        subprocess.run(
            ["curl", "-fL", "--retry", "3", "--retry-delay", "1", url, "-o", str(target)],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    return target


def materialize_inputs(app_dir: Path) -> None:
    app_dir.mkdir(parents=True, exist_ok=True)
    for item in TASK_SPEC.get("inputs", []):
        destination = app_dir / item["dest"]
        destination.parent.mkdir(parents=True, exist_ok=True)
        if not destination.exists():
            shutil.copyfile(fetch_cloud_file(item["url"], item["dest"]), destination)


def run_oracle(app_dir: Path) -> None:
    env = os.environ.copy()
    env["APP_DIR"] = str(app_dir)
    subprocess.run(
        ["bash", str(Path(__file__).resolve().parents[1] / "solution" / "solve.sh")],
        check=True,
        env=env,
    )


def materialize_result_sidecars() -> None:
    for sidecar in TASK_SPEC.get("result_sidecars", []):
        if sidecar["type"] != "libreoffice_csv_export":
            continue
        source = localize_app_path(sidecar["source"])
        output = localize_app_path(sidecar["output"])
        if output.exists() or not source.exists():
            continue
        libreoffice = shutil.which("libreoffice")
        if libreoffice is None:
            continue
        output.parent.mkdir(parents=True, exist_ok=True)
        subprocess.run(
            [
                libreoffice,
                "--convert-to",
                sidecar["convert_to"],
                "--outdir",
                str(output.parent),
                str(source),
            ],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )


def _parse_sheet_idx(
    sheet_idx: Union[int, str],
    result: Union[pd.ExcelFile, Workbook, str],
    expected: Union[pd.ExcelFile, Workbook, str],
    result_sheet_names: List[str],
    expected_sheet_names: List[str],
) -> Tuple[Union[pd.ExcelFile, Workbook, str], str]:
    if isinstance(sheet_idx, int):
        try:
            if not result_sheet_names or sheet_idx >= len(result_sheet_names):
                logger.error(
                    f"Sheet index {sheet_idx} out of range. Available sheets: {result_sheet_names}"
                )
                index = ""
            else:
                index = result_sheet_names[sheet_idx]
                logger.debug(f"Sheet index {sheet_idx} resolved to sheet: {index}")
        except Exception as exc:
            logger.error(f"Error resolving sheet index {sheet_idx}: {exc}")
            index = ""
        book = result
    elif sheet_idx.startswith("RI"):
        try:
            index = result_sheet_names[int(sheet_idx[2:])]
        except Exception:
            index = ""
        book = result
    elif sheet_idx.startswith("RN"):
        index = sheet_idx[2:]
        book = result
    elif sheet_idx.startswith("EI"):
        try:
            index = expected_sheet_names[int(sheet_idx[2:])]
        except Exception:
            index = ""
        book = expected
    elif sheet_idx.startswith("EN"):
        index = sheet_idx[2:]
        book = expected
    else:
        logger.error("Unrecognized sheet index")
        raise ValueError("Unrecognized sheet index")
    return book, index


def _load_sheet(book, index):
    try:
        if isinstance(book, str):
            csv_name = "{:}-{:}.csv".format(os.path.splitext(book)[0], index)
            try:
                all_lines = _safe_read_file(csv_name)
                csv_lines = list(
                    itertools.dropwhile(
                        lambda line: len(line) == 0,
                        map(lambda line: line.strip(), reversed(all_lines)),
                    )
                )
                return csv_lines
            except (FileNotFoundError, IOError) as exc:
                logger.error(f"Failed to read CSV file {csv_name}: {exc}")
                return None
        if isinstance(book, pd.ExcelFile):
            return pd.read_excel(book, index)
        if isinstance(book, Workbook):
            return book[index]
        logger.error("Not supported workbook format")
        raise NotImplementedError("Not supported workbook format")
    except NotImplementedError:
        raise
    except Exception:
        return None


def _freeze_compare_key(sheet: Worksheet) -> Tuple[Any, ...]:
    pane = sheet.sheet_view.pane
    if pane is None:
        return ("none",)
    state = (pane.state or "").lower()
    xs = int(pane.xSplit) if pane.xSplit is not None else 0
    ys = int(pane.ySplit) if pane.ySplit is not None else 0
    if state in ("frozen", "frozensplit"):
        return ("frozen", xs, ys)
    return ("other", xs, ys, pane.topLeftCell)


def _safe_read_file(file_path: str) -> List[str]:
    encodings = [
        "utf-8",
        "utf-8-sig",
        "latin-1",
        "windows-1252",
        "gbk",
        "cp1251",
        "iso-8859-1",
    ]

    last_error = None
    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as handle:
                return handle.read().splitlines()
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except (FileNotFoundError, IOError):
            raise

    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as handle:
            logger.warning(f"Read file {file_path} with UTF-8 and error replacement")
            return handle.read().splitlines()
    except Exception as exc:
        logger.error(
            f"Failed to read file {file_path} with any encoding. Last error: {last_error}"
        )
        raise IOError(f"Cannot read file {file_path} with any supported encoding") from exc


def compare_csv(result: str, expected: Union[str, List[str]], **options) -> float:
    if result is None:
        return 0.0

    try:
        result_lines = _safe_read_file(result)
    except (FileNotFoundError, IOError) as exc:
        logger.error(f"Failed to read result file {result}: {exc}")
        return 0.0

    expected_files = [expected] if isinstance(expected, str) else expected
    for expected_file in expected_files:
        try:
            expected_lines = _safe_read_file(expected_file)
            current_result_lines = result_lines
            current_expected_lines = expected_lines

            if not options.get("strict", True):
                current_result_lines = map(str.strip, current_result_lines)
                current_expected_lines = map(str.strip, current_expected_lines)
            if options.get("ignore_case", False):
                current_result_lines = map(str.lower, current_result_lines)
                current_expected_lines = map(str.lower, current_expected_lines)

            if list(current_result_lines) == list(current_expected_lines):
                return 1.0
        except (FileNotFoundError, IOError):
            continue
    return 0.0


_xlsx_namespaces = [
    ("oo", "http://schemas.openxmlformats.org/spreadsheetml/2006/main"),
    ("x14", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"),
    ("xm", "http://schemas.microsoft.com/office/excel/2006/main"),
]
_xlsx_ns_mapping = dict(_xlsx_namespaces)
_xlsx_ns_imapping = dict(map(lambda item: (item[1], item[0]), _xlsx_namespaces))
_xlsx_ns_imapping["http://schemas.openxmlformats.org/spreadsheetml/2006/main"] = None
_sheet_name_selector = lxml.cssselect.CSSSelector(
    "oo|sheets>oo|sheet", namespaces=_xlsx_ns_mapping
)
_sparklines_selector = lxml.cssselect.CSSSelector(
    "x14|sparkline", namespaces=_xlsx_ns_mapping
)
_shared_str_selector = lxml.cssselect.CSSSelector(
    "oo|sst>oo|si", namespaces=_xlsx_ns_mapping
)
_shared_str_value_selector = lxml.cssselect.CSSSelector(
    "oo|t", namespaces=_xlsx_ns_mapping
)
_absolute_range_pattern: Pattern[str] = re.compile(
    r"""\$(?P<col1>[A-Z]{1,3})\$(?P<row1>\d+)
        (?::\$(?P<col2>[A-Z]{1,3})\$(?P<row2>\d+))?
    """,
    re.X,
)


def load_sparklines(xlsx_file: str, sheet_name: str) -> Dict[str, str]:
    try:
        with zipfile.ZipFile(xlsx_file, "r") as archive:
            with archive.open("xl/workbook.xml") as handle:
                workbook_database = lxml.etree.fromstring(handle.read())
                sheets = _sheet_name_selector(workbook_database)
                sheet_names = {sheet.get("name"): sheet.get("sheetId") for sheet in sheets}
            with archive.open("xl/worksheets/sheet{:}.xml".format(sheet_names[sheet_name])) as handle:
                sheet = lxml.etree.fromstring(handle.read())
                sparklines = _sparklines_selector(sheet)
    except zipfile.BadZipFile:
        return {}

    sparklines_dict = {}
    for sparkline_element in sparklines:
        sparkline_xml = lxml.etree.tostring(sparkline_element, encoding="unicode")
        sparkline = xmltodict.parse(
            sparkline_xml,
            process_namespaces=True,
            namespaces=_xlsx_ns_imapping,
        )
        sparklines_dict[sparkline["x14:sparkline"]["xm:sqref"]] = sparkline["x14:sparkline"]["xm:f"]
    return sparklines_dict


def load_charts(xlsx_file: Workbook, sheet_name: str, **options) -> Dict[str, Any]:
    try:
        worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    charts = worksheet._charts

    chart_set = {}
    chart_props = set(options["chart_props"]) if "chart_props" in options else set()
    for chart in charts:
        series = []
        for item in chart.series:
            if hasattr(item.val, "numRef") and hasattr(item.val.numRef, "f"):
                value_str = item.val.numRef.f
            elif hasattr(item.val, "strRef") and hasattr(item.val.strRef, "f"):
                value_str = item.val.strRef.f
            else:
                value_str = ""
            if hasattr(item.cat, "numRef") and hasattr(item.cat.numRef, "f"):
                categ_str = item.cat.numRef.f
            elif hasattr(item.cat, "strRef") and hasattr(item.cat.strRef, "f"):
                categ_str = item.cat.strRef.f
            else:
                categ_str = ""
            series.append("{:},{:}".format(value_str, categ_str))
        series_key = ";".join(series)

        info = {}
        if "title" in chart_props:
            try:
                info["title"] = chart.title.tx.rich.p[0].r[0].t
            except Exception:
                info["title"] = None
        if "legend" in chart_props:
            info["legend"] = chart.legend.position if chart.legend is not None else None
        if "anchor" in chart_props:
            info["anchor"] = [
                chart.anchor.editAs,
                chart.anchor._from.col,
                chart.anchor.to.row,
                chart.anchor.to.col,
                chart.anchor.to.row,
            ]
        if "width" in chart_props:
            info["width"] = chart.width
        if "height" in chart_props:
            info["height"] = chart.height
        if "type" in chart_props:
            info["type"] = chart.tagname
        if "direction" in chart_props:
            info["direction"] = chart.barDir
        if "xtitle" in chart_props:
            try:
                info["xtitle"] = chart.x_axis.title.tx.rich.p[0].r[0].t
            except Exception:
                info["xtitle"] = None
        if "ytitle" in chart_props:
            try:
                info["ytitle"] = chart.y_axis.title.tx.rich.p[0].r[0].t
            except Exception:
                info["ytitle"] = None
        if "ztitle" in chart_props:
            try:
                info["ztitle"] = chart.z_axis.title.tx.rich.p[0].r[0].t
            except Exception:
                info["ztitle"] = None
        chart_set[series_key] = info
    logger.debug(".[%s].charts: %s", sheet_name, repr(chart_set))
    return chart_set


def load_pivot_tables(xlsx_file: Workbook, sheet_name: str, **options) -> Dict[str, Any]:
    try:
        worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    pivots = worksheet._pivots

    pivot_set = {}
    pivot_props = set(options.get("pivot_props", []))
    for pivot in pivots:
        raw_selection = [
            [(item.h, item.x) for item in field.items if item.x is not None]
            for field in pivot.pivotFields
        ]
        raw_trimmed = list(itertools.dropwhile(lambda row: len(row) == 0, raw_selection))
        left_bias = len(raw_selection) - len(raw_trimmed)
        selection = list(itertools.dropwhile(lambda row: len(row) == 0, reversed(raw_trimmed)))[::-1]
        right_bias = len(raw_trimmed) - len(selection)
        cache_source: PivotCacheSource = pivot.cache.cacheSource
        cell_range1, cell_range2 = cache_source.worksheetSource.ref.split(":")
        coord1 = coordinate_to_tuple(cell_range1)
        coord2 = coordinate_to_tuple(cell_range2)
        coord1 = (coord1[0], coord1[1] + left_bias)
        coord2 = (coord2[0], coord2[1] - right_bias)
        source = "{:};{:}:{:};{:}".format(
            cache_source.type,
            coord1,
            coord2,
            cache_source.worksheetSource.sheet,
        )

        info = {}
        if "name" in pivot_props:
            info["name"] = pivot.name
        if "show_total" in pivot_props:
            info["show_total"] = pivot.visualTotals
        if "show_empty_row" in pivot_props:
            info["show_empty_row"] = pivot.showEmptyRow
        if "show_empty_col" in pivot_props:
            info["show_empty_col"] = pivot.showEmptyCol
        if "show_headers" in pivot_props:
            info["show_headers"] = pivot.showHeaders
        if "location" in pivot_props:
            info["location"] = pivot.location
        if "filter" in pivot_props or "selection" in pivot_props:
            info["selection"] = selection if "ordered" in pivot_props else [set(row) for row in selection]
        if "filter" in pivot_props:
            info["filter_fields"] = set(field.fld for field in pivot.pageFields)
        if "col_fields" in pivot_props:
            info["col_fields"] = [field.x - left_bias for field in pivot.colFields]
        if "row_fields" in pivot_props:
            info["row_fields"] = [field.x - left_bias for field in pivot.rowFields]
        if "data_fields" in pivot_props:
            info["data_fields"] = [
                "{:d};{:};{:};{:}".format(
                    field.fld - left_bias,
                    field.name if "data_fields_name" in pivot_props else "",
                    field.subtotal,
                    field.showDataAs,
                )
                for field in pivot.dataFields
            ]
        pivot_set[source] = info
    logger.debug(".[%s].pivots: %s", sheet_name, repr(pivot_set))
    return pivot_set


def read_cell_value(xlsx_file: str, sheet_name: str, coordinate: str) -> Any:
    logger.debug(
        "Reading cell value from %s, sheet: %s, coordinate: %s",
        xlsx_file,
        sheet_name,
        coordinate,
    )
    if not os.path.exists(xlsx_file):
        logger.error(f"Excel file not found: {xlsx_file}")
        return None

    try:
        with zipfile.ZipFile(xlsx_file, "r") as archive:
            try:
                with archive.open("xl/sharedStrings.xml") as handle:
                    shared_str_xml = lxml.etree.fromstring(handle.read())
                    str_elements = _shared_str_selector(shared_str_xml)
                    shared_strs = [
                        "".join(text.text for text in _shared_str_value_selector(element))
                        for element in str_elements
                    ]
            except Exception:
                shared_strs = []

            with archive.open("xl/workbook.xml") as handle:
                workbook_database = lxml.etree.fromstring(handle.read())
                sheets = _sheet_name_selector(workbook_database)
                sheet_names = {sheet.get("name"): sheet.get("sheetId") for sheet in sheets}

            with archive.open("xl/worksheets/sheet{:}.xml".format(sheet_names[sheet_name])) as handle:
                sheet = lxml.etree.fromstring(handle.read())
                cells = lxml.cssselect.CSSSelector(
                    'oo|row>oo|c[r="{:}"]'.format(coordinate),
                    namespaces=_xlsx_ns_mapping,
                )(sheet)
                if len(cells) == 0:
                    return None
                cell = cells[0]
    except zipfile.BadZipFile as exc:
        logger.error(f"Bad zip file {xlsx_file}: {exc}")
        return None
    except KeyError as exc:
        logger.error(f"Sheet {sheet_name} not found in {xlsx_file}: {exc}")
        return None
    except Exception as exc:
        logger.error(f"Error reading {xlsx_file}: {exc}")
        return None

    cell = xmltodict.parse(
        lxml.etree.tostring(cell, encoding="unicode"),
        process_namespaces=True,
        namespaces=_xlsx_ns_imapping,
    )
    try:
        if "@t" not in cell["c"] or cell["c"]["@t"] == "n":
            return float(cell["c"]["v"])
        if cell["c"]["@t"] == "s":
            return shared_strs[int(cell["c"]["v"])]
        if cell["c"]["@t"] == "str":
            return cell["c"]["v"]
        if cell["c"]["@t"] == "inlineStr":
            return cell["c"]["is"]["t"]
        if cell["c"]["@t"] == "e":
            return cell["c"]["v"]
    except (KeyError, ValueError):
        return None
    return None


def _read_cell_style(style_name: str, cell: Union[Cell, MergedCell], diff_style: Optional[DifferentialStyle] = None) -> Any:
    if style_name == "number_format":
        return (cell.number_format if diff_style is None else diff_style.numFmt.formatCode) if cell.value is not None and cell.data_type == "n" else None
    if style_name == "font_name":
        return (diff_style or cell).font.name if cell.value is not None else None
    if style_name == "font_family":
        return (diff_style or cell).font.family if cell.value is not None else None
    if style_name == "font_color":
        return (diff_style or cell).font.color.rgb if cell.value is not None else None
    if style_name == "font_bold":
        return (diff_style or cell).font.bold if cell.value is not None else None
    if style_name == "font_italic":
        return (diff_style or cell).font.italic if cell.value is not None else None
    if style_name == "font_underline":
        return (diff_style or cell).font.underline if cell.value is not None else None
    if style_name == "font_size":
        return (diff_style or cell).font.size if cell.value is not None else None
    if style_name == "fill_type":
        try:
            return (diff_style or cell).fill.tagname
        except Exception:
            return None
    if style_name in {"bgcolor", "fgcolor"}:
        try:
            if diff_style is not None:
                return diff_style.fill.bgColor.rgb
            return cell.fill.fgColor.rgb
        except Exception:
            return None
    if style_name == "hyperlink":
        return cell.hyperlink or "" if cell.value is not None else None
    if style_name == "merge":
        return isinstance(cell, MergedCell)
    raise NotImplementedError("Unsupported Style: {:}".format(style_name))


def _process_xlsx_cf_operator(operator_name: str, value: Any, ref: List[Any]) -> bool:
    try:
        if operator_name == "lessThanOrEqual":
            result = value <= ref[0]
        elif operator_name == "lessThan":
            result = value < ref[0]
        elif operator_name == "equal":
            result = value == ref[0]
        elif operator_name == "greaterThanOrEqual":
            result = value >= ref[0]
        elif operator_name == "notEqual":
            result = value != ref[0]
        elif operator_name == "greaterThan":
            result = value > ref[0]
        elif operator_name == "between":
            small_one, large_one = min(ref), max(ref)
            result = value >= small_one and value <= large_one
        elif operator_name == "notBetween":
            small_one, large_one = min(ref), max(ref)
            result = value < small_one or value > large_one
        else:
            logger.exception("Not Implemented CondFormat Operator: {:}".format(operator_name))
            return False
        return result
    except TypeError:
        logger.exception("Unmatched type of %s and %s. Auto to False", repr(value), repr(ref))
        return False
    except IndexError:
        logger.exception("ref array doesn't have enough elements. Auto to False: %s", repr(ref))
        return False


def load_xlsx_styles(xlsx_file: Workbook, sheet_name: str, book_name: str, **options) -> Dict[str, List[Any]]:
    try:
        worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}

    style_dict = {}
    concerned_styles = options.get("props", [])

    for column in worksheet.iter_cols():
        for cell in column:
            style_dict[cell.coordinate] = [_read_cell_style(style_name, cell) for style_name in concerned_styles]

    conditional_formattings: ConditionalFormattingList = worksheet.conditional_formatting
    formula_parser = formulas.Parser()
    for formatting in conditional_formattings:
        for rule in formatting.rules:
            active_cells = []
            formulae = []
            argument_lists = []
            has_error = False
            for formula_text in rule.formula:
                try:
                    formula_func = formula_parser.ast("=" + formula_text)[1].compile()
                except Exception:
                    logger.exception("Formula parsing error: %s. Skipping.", repr(formula_text))
                    has_error = True
                    break

                arguments = []
                absolute_range_match = _absolute_range_pattern.findall(formula_text)
                for match in absolute_range_match:
                    if match[2] is None and match[3] is None:
                        arguments.append(read_cell_value(book_name, sheet_name, coordinate="{:}{:}".format(match[0], match[1])))
                    else:
                        arguments.append([
                            read_cell_value(
                                book_name,
                                sheet_name,
                                coordinate="{:}{:}".format(get_column_letter(coord[1]), coord[0]),
                            )
                            for coord in CellRange("{:}{:}:{:}{:}".format(match[0], match[1], match[2], match[3])).cells
                        ])
                formulae.append(formula_func)
                argument_lists.append(arguments)

            if has_error:
                continue

            if rule.type in {"expression", "containsText", "notContainsText", "endsWith", "beginsWith", "containsErrors", "notContainsErrors"}:
                condition = formulae[0]
                arguments = argument_lists[0]
                is_active = lambda value: condition(value, *arguments)
            elif rule.type == "cellIs":
                operator_name = rule.operator
                try:
                    references = [formula() for formula in formulae]
                except Exception:
                    logger.exception("Error occurs while calculating reference values for cellIs condition formatting.")
                    continue
                is_active = lambda value: _process_xlsx_cf_operator(operator_name, value, references)
            else:
                logger.exception("Not Implemented Condition Type: {:}".format(rule.type))
                continue

            nb_contiguous_nothings = 0
            for cell_range in formatting.cells:
                for coords in cell_range.cells:
                    cell = worksheet.cell(row=coords[0], column=coords[1])
                    cell_value = read_cell_value(
                        book_name,
                        sheet_name,
                        coordinate="{:}{:d}".format(get_column_letter(coords[1]), coords[0]),
                    )
                    if cell_value is None:
                        nb_contiguous_nothings += 1
                        if nb_contiguous_nothings > 50:
                            break
                        continue
                    try:
                        satisfies_condition = is_active(cell_value)
                    except Exception:
                        logger.exception("Error in formula calculation with cell value %d", repr(cell_value))
                        satisfies_condition = False
                    if satisfies_condition:
                        active_cells.append(cell)

            for active_cell in active_cells:
                style_dict[active_cell.coordinate] = [
                    _read_cell_style(style_name, active_cell, rule.dxf)
                    for style_name in concerned_styles
                ]

    logger.debug(".[%s].styles: %s", sheet_name, repr(style_dict))
    return style_dict


def load_rows_or_cols(xlsx_file: Workbook, sheet_name: str, **options) -> Dict[Union[int, str], Dict[str, Any]]:
    try:
        worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    objs: DimensionHolder = getattr(worksheet, "{:}_dimensions".format(options["obj"]))

    obj_set = {}
    obj_props = set(options.get("props", []))
    for obj_no, obj_dims in objs.items():
        info_dict = {}
        for prop in obj_props:
            info_dict[prop] = getattr(obj_dims, prop)
        obj_set[obj_no] = info_dict
    return obj_set


def load_filters(xlsx_file: Workbook, sheet_name: str, **options) -> Dict[str, Any]:
    try:
        worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}

    filters: AutoFilter = worksheet.auto_filter
    filter_dict = {"ref": filters.ref}

    filter_column_set = []
    for filter_column in filters.filterColumn:
        column = {
            "col_id": filter_column.colId,
            "hidden_button": filter_column.hiddenButton,
            "show_button": filter_column.showButton,
        }
        if filter_column.filters is not None:
            column["filters_blank"] = filter_column.filters.blank
            column["filters"] = set(filter_column.filters.filter)
        if filter_column.customFilters is not None:
            column["custom_filters_op"] = filter_column.customFilters._and
            column["custom_filters"] = set(
                (flt.operator, flt.val) for flt in filter_column.customFilters.customFilter
            )
        filter_column_set.append(column)
    filter_dict["filter_column"] = list(sorted(filter_column_set, key=lambda item: item["col_id"]))

    sort_state: Optional[SortState] = filters.sortState
    if sort_state is not None:
        filter_dict["sort_state"] = {
            "sort": sort_state.columnSort,
            "case": sort_state.caseSensitive,
            "method": sort_state.sortMethod,
            "ref": sort_state.ref,
            "condition": [
                {
                    "descending": condition.descending,
                    "key": condition.sortBy,
                    "ref": condition.ref,
                    "custom_list": condition.customList,
                    "dxf_id": condition.dxfId,
                    "icon": condition.iconSet,
                    "iconid": condition.iconId,
                }
                for condition in sort_state.sortCondition
            ],
        }
    return filter_dict


def _multicellrange_containsby(subset_candidate: MultiCellRange, superset_candidate: MultiCellRange) -> bool:
    return all(cell_range in superset_candidate for cell_range in subset_candidate)


def _match_value_to_rule(value: V, rule: Dict[str, Union[str, V]]) -> bool:
    if rule["method"].startswith("re"):
        flags = (getattr(re, flag_name) for flag_name in rule["method"].split(".")[1:])
        flag = functools.reduce(operator.or_, flags, re.RegexFlag(0))
        return re.search(rule["ref"], value, flag) is not None
    if rule["method"] in {"eq", "ne", "le", "lt", "ge", "gt"}:
        return getattr(operator, rule["method"])(value, rule["ref"])
    if rule["method"].startswith("approx"):
        threshold = float(rule["method"].split(":")[1])
        try:
            value = float(value)
        except (ValueError, TypeError):
            return False
        return abs(value - rule["ref"]) <= threshold
    if rule["method"] == "spreadsheet_range":
        subset_limit = MultiCellRange(rule["ref"][0])
        superset_limit = MultiCellRange(rule["ref"][1])
        return _multicellrange_containsby(subset_limit, value) and _multicellrange_containsby(value, superset_limit)
    if rule["method"].startswith("range."):
        left_et = rule["method"][6]
        right_et = rule["method"][7]
        return getattr(operator, "l" + left_et)(rule["ref"][0], value) and getattr(operator, "l" + right_et)(value, rule["ref"][1])
    if rule["method"] in {"str_list_eq", "str_set_eq"}:
        container_type_str = rule["method"][4:-3]
        container_type = getattr(__import__("builtins"), container_type_str)
        value = container_type(value.strip("\"'").split(","))
        ref = container_type(rule["ref"])
        return value == ref
    raise NotImplementedError()


def compare_table(result: str, expected: str = None, **options) -> float:
    if result is None:
        logger.error("Result file path is None")
        return 0.0
    if not os.path.exists(result):
        logger.error(f"Result file not found: {result}")
        return 0.0

    try:
        xlworkbookr: Workbook = openpyxl.load_workbook(filename=result)
        pdworkbookr = pd.ExcelFile(result)
    except Exception as exc:
        logger.error(f"Failed to load result file {result}: {exc}")
        return 0.0
    worksheetr_names = pdworkbookr.sheet_names

    if expected is not None:
        xlworkbooke = openpyxl.load_workbook(filename=expected)
        pdworkbooke = pd.ExcelFile(expected)
        worksheete_names = pdworkbooke.sheet_names
    else:
        xlworkbooke = None
        pdworkbooke = None
        worksheete_names = None

    parse_idx = functools.partial(
        _parse_sheet_idx,
        result_sheet_names=worksheetr_names,
        expected_sheet_names=worksheete_names,
    )

    passes = True
    for rule in options["rules"]:
        if rule["type"] == "sheet_name":
            metric = worksheetr_names == worksheete_names

        elif rule["type"] == "sheet_data":
            error_limit = rule.get("precision", 4)
            sheet1 = _load_sheet(*parse_idx(rule["sheet_idx0"], pdworkbookr, pdworkbooke))
            if sheet1 is None:
                return 0.0
            sheet2 = _load_sheet(*parse_idx(rule["sheet_idx1"], pdworkbookr, pdworkbooke))
            sheet1 = sheet1.round(error_limit)
            sheet2 = sheet2.round(error_limit)
            metric = sheet1.equals(sheet2)

        elif rule["type"] == "sheet_print":
            sheet1 = _load_sheet(*parse_idx(rule["sheet_idx0"], result, expected))
            if sheet1 is None:
                return 0.0
            sheet2 = _load_sheet(*parse_idx(rule["sheet_idx1"], result, expected))
            if rule.get("ignore_case", False):
                sheet1 = [line.lower() for line in sheet1]
                sheet2 = [line.lower() for line in sheet2]
            metric = sheet1 == sheet2

        elif rule["type"] == "sheet_fuzzy":
            sheet1 = parse_idx(rule["sheet_idx0"], result, expected)
            sheet2 = parse_idx(rule["sheet_idx1"], result, expected)
            total_metric = True
            for range_rule in rule["rules"]:
                for range_item in MultiCellRange(range_rule["range"]):
                    for cell in range_item.cells:
                        coordinate = "{:}{:d}".format(get_column_letter(cell[1]), cell[0])
                        value1 = str(read_cell_value(*sheet1, coordinate))
                        value2 = str(read_cell_value(*sheet2, coordinate))

                        for replacement in range_rule.get("normalization", []):
                            value1 = value1.replace(replacement[0], replacement[1])
                            value2 = value2.replace(replacement[0], replacement[1])
                        if "trim_leadings" in range_rule:
                            value1 = value1.lstrip(range_rule["trim_leadings"])
                            value2 = value2.lstrip(range_rule["trim_leadings"])
                        if "trim_trailings" in range_rule:
                            value1 = value1.rstrip(range_rule["trim_trailings"])
                            value2 = value2.rstrip(range_rule["trim_trailings"])
                        if "ignore_chars" in range_rule:
                            ignore_chars = set(range_rule["ignore_chars"])
                            value1 = "".join(filter(lambda char: char not in ignore_chars, value1))
                            value2 = "".join(filter(lambda char: char not in ignore_chars, value2))
                        if range_rule.get("ignore_case", False):
                            value1 = value1.lower()
                            value2 = value2.lower()

                        if range_rule["type"] == "includes":
                            metric = value2 in value1
                        elif range_rule["type"] == "included_by":
                            metric = value1 in value2
                        elif range_rule["type"] == "fuzzy_match":
                            metric = fuzz.ratio(value1, value2) >= range_rule.get("threshold", 85.0)
                        elif range_rule["type"] == "exact_match":
                            metric = value1 == value2
                        total_metric = total_metric and metric
            metric = total_metric

        elif rule["type"] == "sparkline":
            sparkline1 = load_sparklines(*parse_idx(rule["sheet_idx0"], result, expected))
            sparkline2 = load_sparklines(*parse_idx(rule["sheet_idx1"], result, expected))
            metric = sparkline1 == sparkline2

        elif rule["type"] == "chart":
            charts1 = load_charts(*parse_idx(rule["sheet_idx0"], xlworkbookr, xlworkbooke), **rule)
            charts2 = load_charts(*parse_idx(rule["sheet_idx1"], xlworkbookr, xlworkbooke), **rule)
            metric = charts1 == charts2

        elif rule["type"] == "style":
            sheet_idx1 = parse_idx(rule["sheet_idx0"], xlworkbookr, xlworkbooke)
            book_name1 = parse_idx(rule["sheet_idx0"], result, expected)[0]
            styles1 = load_xlsx_styles(*sheet_idx1, book_name1, **rule)

            sheet_idx2 = parse_idx(rule["sheet_idx1"], xlworkbookr, xlworkbooke)
            book_name2 = parse_idx(rule["sheet_idx1"], result, expected)[0]
            styles2 = load_xlsx_styles(*sheet_idx2, book_name2, **rule)
            metric = styles1 == styles2

        elif rule["type"] == "freeze":
            sheet1 = _load_sheet(*parse_idx(rule["sheet_idx0"], xlworkbookr, xlworkbooke))
            if sheet1 is None:
                return 0.0
            sheet2 = _load_sheet(*parse_idx(rule["sheet_idx1"], xlworkbookr, xlworkbooke))
            metric = _freeze_compare_key(sheet1) == _freeze_compare_key(sheet2)

        elif rule["type"] == "zoom":
            sheet = _load_sheet(*parse_idx(rule["sheet_idx"], xlworkbookr, xlworkbooke))
            if sheet is None:
                return 0.0
            zoom_scale = sheet.sheet_view.zoomScale or 100.0
            metric = _match_value_to_rule(zoom_scale, rule)

        elif rule["type"] == "data_validation":
            sheet = _load_sheet(*parse_idx(rule["sheet_idx"], xlworkbookr, xlworkbooke))
            if sheet is None:
                return 0.0
            data_validators = sheet.data_validations.dataValidation
            total_metric = len(data_validators) >= len(rule["dv_props"])
            for validator in data_validators:
                metric = False
                for props in rule["dv_props"]:
                    metric = metric or all(
                        _match_value_to_rule(getattr(validator, attr), matcher)
                        for attr, matcher in props.items()
                    )
                    if metric:
                        break
                total_metric = total_metric and metric
                if not total_metric:
                    break
            metric = total_metric

        elif rule["type"] == "row_props":
            rows1 = load_rows_or_cols(*parse_idx(rule["sheet_idx0"], xlworkbookr, xlworkbooke), obj="row", **rule)
            rows2 = load_rows_or_cols(*parse_idx(rule["sheet_idx1"], xlworkbookr, xlworkbooke), obj="row", **rule)
            metric = rows1 == rows2

        elif rule["type"] == "col_props":
            cols1 = load_rows_or_cols(*parse_idx(rule["sheet_idx0"], xlworkbookr, xlworkbooke), obj="column", **rule)
            cols2 = load_rows_or_cols(*parse_idx(rule["sheet_idx1"], xlworkbookr, xlworkbooke), obj="column", **rule)
            metric = cols1 == cols2

        elif rule["type"] == "filter":
            filters1 = load_filters(*parse_idx(rule["sheet_idx0"], xlworkbookr, xlworkbooke), **rule)
            filters2 = load_filters(*parse_idx(rule["sheet_idx1"], xlworkbookr, xlworkbooke), **rule)
            metric = filters1 == filters2

        elif rule["type"] == "pivot_table":
            pivots1 = load_pivot_tables(*parse_idx(rule["sheet_idx0"], xlworkbookr, xlworkbooke), **rule)
            pivots2 = load_pivot_tables(*parse_idx(rule["sheet_idx1"], xlworkbookr, xlworkbooke), **rule)
            metric = pivots1 == pivots2

        elif rule["type"] == "check_cell":
            try:
                sheet = _load_sheet(*parse_idx(rule["sheet_idx"], xlworkbookr, xlworkbooke))
                if sheet is None:
                    return 0.0
                cell = sheet[rule["coordinate"]]
                metric = True
                for prop, matcher in rule["props"].items():
                    if prop == "value":
                        try:
                            parsed_result = parse_idx(rule["sheet_idx"], result, expected)
                            value = read_cell_value(*parsed_result, rule["coordinate"])
                        except Exception as exc:
                            logger.error(f"Failed to read cell value at {rule['coordinate']}: {exc}")
                            value = None
                    else:
                        try:
                            value = _read_cell_style(prop, cell)
                        except Exception as exc:
                            logger.error(f"Failed to read cell style {prop} at {rule['coordinate']}: {exc}")
                            value = None
                    metric = metric and _match_value_to_rule(value, matcher)
            except Exception as exc:
                logger.error(f"Error in check_cell processing: {exc}")
                return 0.0

        else:
            raise NotImplementedError("Unimplemented sheet check: {:}".format(rule["type"]))

        passes = passes and metric
        if not passes:
            break

    return float(passes)


def check_pdf_pages(pdf_file: str, rules: Dict[str, Any]) -> float:
    if pdf_file is None:
        return 0.0
    reader = PdfReader(pdf_file)
    nb_pages = len(reader.pages)
    return float(getattr(operator, rules["relation"])(nb_pages, rules["ref_value"]))


def compare_pdfs(pdf1_path: Union[str, List[str]], pdf2_path: Union[str, List[str]]):
    if type(pdf2_path) != list:
        pdf1_path, pdf2_path = [pdf1_path], [pdf2_path]

    def extract_text_from_pdf(pdf_path):
        text = ""
        with fitz.open(pdf_path) as pdf:
            for page in pdf:
                text += page.get_text()
        return text.strip()

    score = 0.0
    for path1, path2 in zip(pdf1_path, pdf2_path):
        try:
            text1 = extract_text_from_pdf(path1)
            text2 = extract_text_from_pdf(path2)
            score += fuzz.ratio(text1, text2) / 100
        except Exception as exc:
            logger.info(f"[ERROR]: unexpected error occurred when comparing PDF files: {exc}")
    return score / len(pdf2_path)


def check_infeasible_response(path: Path, required_terms: List[str]) -> float:
    if path is None or not path.exists():
        return 0.0
    text = path.read_text(encoding="utf-8").strip()
    if not text:
        return 0.0
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines or lines[0].lower() != "infeasible":
        return 0.0
    reason = " ".join(lines[1:]).lower()
    if not reason:
        return 0.0
    if not all(term in reason for term in required_terms):
        return 0.0
    if not any(phrase in reason for phrase in LIMITATION_PHRASES):
        return 0.0
    return 1.0


def resolve_expected(spec):
    if spec is None:
        return None
    if spec["type"] == "cloud_file":
        return str(fetch_cloud_file(spec["url"], spec["dest"]))
    if spec["type"] == "cloud_file_bundle":
        primary = spec["primary"]
        primary_path = fetch_cloud_file(primary["url"], primary["dest"])
        for sidecar in spec.get("sidecars", []):
            fetch_cloud_file(sidecar["url"], sidecar["dest"])
        return str(primary_path)
    if spec["type"] == "cloud_file_list":
        return [str(fetch_cloud_file(url, dest)) for url, dest in zip(spec["urls"], spec["dests"])]
    if spec["type"] == "vm_file":
        return str(localize_app_path(spec["path"]))
    if spec["type"] == "rule":
        return spec["rules"]
    raise ValueError(f"Unsupported expected spec: {spec}")


def resolve_result_state(result_state):
    if isinstance(result_state, list):
        return [resolve_result_state(item) for item in result_state]
    if isinstance(result_state, str) and result_state.startswith("/app/"):
        return str(localize_app_path(result_state))
    return result_state


def evaluate_single(func_name, result_state, expected_state=None, options=None):
    options = options or {}
    metric = {
        "compare_table": compare_table,
        "compare_csv": compare_csv,
        "check_pdf_pages": check_pdf_pages,
        "compare_pdfs": compare_pdfs,
    }[func_name]
    try:
        if expected_state is None:
            return float(metric(result_state, **options))
        return float(metric(result_state, expected_state, **options))
    except Exception as exc:
        logger.error("Metric %s failed: %s", func_name, exc)
        return 0.0


def run_evaluation():
    metric_spec = TASK_SPEC["metric_spec"]
    materialize_result_sidecars()
    if metric_spec.get("mode") == "infeasible":
        return check_infeasible_response(localize_app_path(TASK_SPEC["output_path"]), metric_spec["required_terms"])

    func = metric_spec["func"]
    if isinstance(func, list):
        conj = metric_spec.get("conj", "and")
        results = []
        for index, func_name in enumerate(func):
            result_state = resolve_result_state(metric_spec["result"][index])
            expected_state = resolve_expected(metric_spec["expected"][index]) if "expected" in metric_spec else None
            options = metric_spec.get("options", [{}] * len(func))[index]
            metric_value = evaluate_single(func_name, result_state, expected_state, options)
            if conj == "and" and metric_value == 0.0:
                return 0.0
            if conj == "or" and metric_value == 1.0:
                return 1.0
            results.append(metric_value)
        return sum(results) / len(results) if conj == "and" else max(results)

    result_state = resolve_result_state(metric_spec["result"])
    expected_state = resolve_expected(metric_spec["expected"]) if "expected" in metric_spec else None
    return evaluate_single(func, result_state, expected_state, metric_spec.get("options", {}))


def run_sanity():
    with tempfile.TemporaryDirectory(prefix=f"{TASK_SPEC['slug']}-") as temp_dir:
        app_dir = Path(temp_dir)
        materialize_inputs(app_dir)
        previous = os.environ.get("APP_DIR")
        os.environ["APP_DIR"] = str(app_dir)
        try:
            fail_score = run_evaluation()
            run_oracle(app_dir)
            pass_score = run_evaluation()
        finally:
            if previous is None:
                os.environ.pop("APP_DIR", None)
            else:
                os.environ["APP_DIR"] = previous
    return {"fail_score": fail_score, "pass_score": pass_score}


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--sanity":
        print(json.dumps(run_sanity(), sort_keys=True))
    else:
        print(run_evaluation())
